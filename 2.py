import sys
from PyQt5 import QtWidgets, uic
from openpyxl import load_workbook

class Qestions(QtWidgets.QWidget):
    def __init__(self, ui, basa, mViget):
        super(Qestions, self).__init__()
        uic.loadUi(ui, self)
        self.wb = load_workbook(basa)
        self.mViget = mViget
        sheet = self.wb[self.wb.sheetnames[0]]
        variant_1 = list(map(int, sheet['C2'].value.split(';')))
        variant_2 = list(map(int, sheet['C3'].value.split(';')))
        self.variants = [variant_1, variant_2]
        self.progression = 0
        self.result = {
    tuple(list(map(int, sheet['D2'].value.split('-')))): sheet['E2'].value,
    tuple(list(map(int, sheet['D3'].value.split('-')))): sheet['E3'].value,
    tuple(list(map(int, sheet['D4'].value.split('-')))): sheet['E4'].value,
    tuple(list(map(int, sheet['D5'].value.split('-')))): sheet['E5'].value,
    tuple(list(map(int, sheet['D6'].value.split('-')))): sheet['E6'].value,
}
        for num in range(1, 1000000):
            if sheet[f'A{num}'].value is None:
                self.max_border = num
                break
        self.questions = [sheet[f'A{val}'].value for val in range(2, self.max_border)]
        self.answers = [sheet[f'B{val}'].value for val in range(2, self.max_border)]
        self.iresult = "вы ещё не прошли тест"

        self.groupOfAnses = QtWidgets.QButtonGroup()
        self.num = 0
        self.otvet.clicked.connect(self.chekAnses)
        self.groupOfAnses.addButton(self.b1)
        self.groupOfAnses.addButton(self.b2)
        self.groupOfAnses.addButton(self.b3)
        self.groupOfAnses.addButton(self.b4)
        self.groupOfAnses.addButton(self.b5)
        self.qst.setText(self.questions[self.num])

        #эту шнягу можно быдет подключить к меню, когда оно будет
        self.menu.clicked.connect(self.qstExit)

    #проверить ответы
    def chekAnses(self):
        btn =self.groupOfAnses.checkedButton()
        if btn != None:
            variable_num = self.answers[self.num]
            self.progression += self.variants[variable_num - 1][abs(self.groupOfAnses.checkedId())-2]
            self.nqst()

    # выдать результат всего тестирования
    def getResult(self):
        for find_result in self.result:
            if find_result[0] <= self.progression <= find_result[1]:
                self.iresult = self.result[find_result]
                print(self.iresult)
                break

    # проверка на есть ли еше вопросы
    def nqst(self):
        self.num += 1
        if self.num >= len(self.questions):
            self.getResult()
            self.hide()
            return
        self.qst.setText(self.questions[self.num])

    def qstExit(self):
        self.mViget.setSelf()
        self.mViget.result = self.iresult
        self.hide()





if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    prog = Qestions("untitled.ui","./database_test.xlsx")
    prog.show()
    sys.exit(app.exec())