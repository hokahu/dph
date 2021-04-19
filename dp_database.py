from openpyxl import load_workbook

# ИЗВЛЕЧЕНИЕ ДАННЫХ И СОЗДАНИЕ ПЕРЕМЕННЫХ ДЛЯ ИХ ХРАНЕНИЯ
# ______________________________________________________________

# Таблица эксель и ее страница
wb = load_workbook('./database_test.xlsx')
sheet = wb[wb.sheetnames[0]]

# Переменные с возможными баллами за ответ
variant_1 = list(map(int, sheet['C2'].value.split(';')))
variant_2 = list(map(int, sheet['C3'].value.split(';')))
variants = [variant_1, variant_2]
# Шкала прогресса, в которую будут складываться все очки
progression = 0

# Возможные итоговые результаты. Ключ - радиус между баллами,
# значение - строка для вывода
result = {
    tuple(list(map(int, sheet['D2'].value.split('-')))): sheet['E2'].value,
    tuple(list(map(int, sheet['D3'].value.split('-')))): sheet['E3'].value,
    tuple(list(map(int, sheet['D4'].value.split('-')))): sheet['E4'].value,
    tuple(list(map(int, sheet['D5'].value.split('-')))): sheet['E5'].value,
    tuple(list(map(int, sheet['D6'].value.split('-')))): sheet['E6'].value,
}

# Определение количества вопросов в стоблце А
# Соответсвенно, сколько вопросов, столько и ответов в стоблце B
for num in range(1, 1000000):
    if sheet[f'A{num}'].value is None:
        max_border = num
        break

# Списки с вопросами и вариациями ответа
questions = [sheet[f'A{val}'].value for val in range(2, max_border)]
print(questions)
answers = [sheet[f'B{val}'].value for val in range(2, max_border)]


possible_answers = [
    'Да', 'Скорее да, чем нет', 'Не знаю',
    'Скорее нет, чем да', 'Нет'
                    ]

# ______________________________________________________________

# ТЕСТИРОВАНИЕ
# ______________________________________________________________

# Проходит по вопросам и задает их пользователю,
# а также сверяет ответы с вариациями


def test():
    global progression
    for question in range(0, max_border - 2):
        print(f'Вопрос {question + 1}')
        # Задает вопрос до тех пор, пока пользователь не
        # внесет корректный ответ
        while True:
            print(questions[question])
            print('Выберите вариант ответа: ')
            for answer in range(len(possible_answers)):
                print(f'{answer + 1}: {possible_answers[answer]}', end='\n')
            user_answer = input()
            if not user_answer.isdigit():
                print('Введите число')
                continue
            user_answer = int(user_answer)
            if user_answer - 1 not in range(len(possible_answers)):
                print('Такого варианта не существует. Выберите снова\n\n')
            else:
                break
        variable_num = answers[question]
        progression += variants[variable_num - 1][user_answer - 1]
    # Сравнивает баллы пользователя с промежутками в итоговых
    # результатах и выводит их
    for find_result in result:
        if find_result[0] <= progression <= find_result[1]:
            print(result[find_result])
            break
test()
# ______________________________________________________________
