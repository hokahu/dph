import sys
from PyQt5 import QtWidgets, uic
from openpyxl import load_workbook


class Qestions(QtWidgets.QWidget):
    def __init__(self, ui, basa, mViget):
        super(Qestions, self).__init__()
        uic.loadUi(ui, self)
        self.wb = load_workbook(basa)
        self.mViget = mViget
        sheet = self.wb[self.wb.sheetnames[0]]
        variant_1 = list(map(int, sheet['C2'].value.split(';')))
        variant_2 = list(map(int, sheet['C3'].value.split(';')))
        self.variants = [variant_1, variant_2]
        self.progression = 0
        self.result = {
    tuple(list(map(int, sheet['D2'].value.split('-')))): sheet['E2'].value,
    tuple(list(map(int, sheet['D3'].value.split('-')))): sheet['E3'].value,
    tuple(list(map(int, sheet['D4'].value.split('-')))): sheet['E4'].value,
    tuple(list(map(int, sheet['D5'].value.split('-')))): sheet['E5'].value,
    tuple(list(map(int, sheet['D6'].value.split('-')))): sheet['E6'].value,
}
        for num in range(1, 1000000):
            if sheet[f'A{num}'].value is None:
                self.max_border = num
                break
        self.questions = [sheet[f'A{val}'].value for val in range(2, self.max_border)]
        self.answers = [sheet[f'B{val}'].value for val in range(2, self.max_border)]
        self.iresult = "вы ещё не прошли тест"

        self.groupOfAnses = QtWidgets.QButtonGroup()
        self.num = 0
        self.otvet.clicked.connect(self.chekAnses)
        self.groupOfAnses.addButton(self.b1)
        self.groupOfAnses.addButton(self.b2)
        self.groupOfAnses.addButton(self.b3)
        self.groupOfAnses.addButton(self.b4)
        self.groupOfAnses.addButton(self.b5)
        self.qst.setText(self.questions[self.num])

        #эту шнягу можно быдет подключить к меню, когда оно будет
        self.menu.clicked.connect(self.qstExit)

    #проверить ответы
    def chekAnses(self):
        btn =self.groupOfAnses.checkedButton()
        if btn != None:
            variable_num = self.answers[self.num]
            self.progression += self.variants[variable_num - 1][abs(self.groupOfAnses.checkedId())-2]
            self.nqst()

    # выдать результат всего тестирования
    def getResult(self):
        for find_result in self.result:
            if find_result[0] <= self.progression <= find_result[1]:
                self.iresult = self.result[find_result]
                self.qstExit()
                break

    # проверка на есть ли еше вопросы
    def nqst(self):
        self.num += 1
        if self.num >= len(self.questions):
            self.getResult()
            self.hide()
            return
        self.qst.setText(self.questions[self.num])

    def qstExit(self):
        self.mViget.setSelf()


class Advice(QtWidgets.QWidget):
    def __init__(self, ui, adv, mViget):
        super(Advice, self).__init__()
        uic.loadUi(ui, self)
        self.mViget = mViget
        self.adv = adv
        self.num = 0
        self.nadv(0)
        self.pr.clicked.connect(lambda: self.nadv(1))
        self.pl.clicked.connect(lambda: self.nadv(-1))
        self.menu.clicked.connect(self.adExit)


    def nadv(self, step):
        self.num = (self.num + step) % len(self.adv)
        text = self.adv[self.num]
        self.l1.setText(text)

    def adExit(self):
        self.mViget.setSelf()



class Menu(QtWidgets.QWidget):
    def __init__(self, ui, basa):
        super(Menu, self).__init__()
        self.ui = ui
        uic.loadUi(self.ui[0], self)
        self.basa = basa
        self.widget = QtWidgets.QMainWindow()
        style = """QMainWindow{
        background-color: white;
        }
        """
        self.widget.setStyleSheet(style)
        self.widget.show()
        self.widget.setCentralWidget(self)
        self.result = "вы ещё не прошли тест"

        self.sbtn.clicked.connect(self.soveti)
        self.qbtn.clicked.connect(self.voprosi)

    def soveti(self):
        self.widget.takeCentralWidget()
        if self.result:
            self.result = self.voprosi.iresult
        self.soveti = Advice(self.ui[2], self.result.split(". "), self)
        self.widget.setCentralWidget(self.soveti)

    def voprosi(self):
        self.widget.takeCentralWidget()
        self.voprosi = Qestions(self.ui[1], self.basa, self)
        self.widget.setCentralWidget(self.voprosi)

    def setSelf(self):
        self.widget.takeCentralWidget()
        self.widget.setCentralWidget(self)



if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    #меню, вопросы, советы
    ui = ["./joker1.ui", "./untitled.ui", "./untitled1.ui"]
    prog = Menu(ui,"./database_test.xlsx")
    prog.show()
    sys.exit(app.exec())
